import sys
from PyQt5.QtWidgets import QMainWindow, QApplication
from PyQt5 import uic
from connect import def_conn
from bs4 import BeautifulSoup
import urllib.request
import os
import xlrd
import xlwt
import openpyxl
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from xlutils3.copy import copy

Ui_MainWindow, QtBaseClass = uic.loadUiType("mainwindow.ui")

class MyApp(QMainWindow):
    def __init__(self):
        super(MyApp, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.ui.dwnldButton.clicked.connect(self.download)
        self.ui.parseButton.clicked.connect(self.parse)
        self.ui.updGlButton.clicked.connect(self.updateGrouplist)

    def updateGrouplist(self):
        cursor = cnx.cursor()
        cursor.execute("SELECT name,code,year FROM groups")
        grouplist = cursor.fetchall()
        print(type(grouplist))
        print(grouplist)
        for group in grouplist:
            print(type(group))
            print('-'.join(map(str, group)))
            self.ui.groupComboBox.addItem('-'.join(map(str, group)))

    def download(self):
        print("downloading...")
        html_doc = urllib.request.urlopen('https://www.mirea.ru/education/schedule-main/schedule/').read()
        soup = BeautifulSoup(html_doc, "html.parser")

        for links in soup.find_all('a'):
            if links.get('href').find("IIT-2k-17_18-vesna.xlsx") != -1:
                link = links.get('href')
                print(link)
        urllib.request.urlretrieve(link, "files/1.xlsx")

    def parse(self):
        print("parsing...")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MyApp()
    window.show()
    cnx = def_conn()

    sys.exit(app.exec_())
    cnx.close()



# size =os.path.getsize("files/1.xlsx")
# print(size)
# f=open("files/latest.txt", 'w')
# f.write(str(os.path.getsize("files/1.xlsx")))
# f.close()
# f=open("files/latest.txt", 'r')
# print(f.readline())
# f.close()

#workbook = xlrd.open_workbook("files/1.xlsx")
#sheet = workbook.sheet_by_index(0)
# first_row = []
# for col in range(sheet.ncols):
#     first_row.append(sheet.cell_value(2,col) )
# # tronsform the workbook to a list of dictionnary
# data =[]
# for row in range(1, sheet.nrows):
#     elm = {}
#     for col in range(sheet.ncols):
#         elm[first_row[col]]=sheet.cell_value(row,col)
#     data.append(elm)
# print(data)

#x=0
#for rows in range(170):
#    if sheet.row_values(2)[rows]=="ИКБО-06-16":
#        print(rows)
#        x=rows
    #print(sheet.row_values(2)[rows])
#vals = [sheet.row_values(rownum)[x] for rownum in range(73)]
#print(vals)

# wb = xlwt.Workbook()
# ws = wb.add_sheet("1")
# i=0
# for rec in vals[0]:
#     ws.write(i,1,rec[0])
#     i =+ i
# wb.save("files/2.xls")
# wb = load_workbook(filename="files/1.xlsx", read_only=True)
# #vals = [v[0].value for v in sheet.range('DY:EB')]
# ws = wb['Лист1']
# for cell in ws.rows:
#     print(cell[128].value)
